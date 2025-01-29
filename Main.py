import cv2 
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os
import numpy as np
from pdf2image import convert_from_path
import re
from nltk.tokenize import sent_tokenize

file_path = r"D:\.vscode\Python\csv\invoice.png"

def load_input(file_path):
    images = []
    try:
        if file_path.lower().endswith('.pdf'):
            images = convert_from_path(file_path)
            images = [np.array(img) for img in images]
        elif file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif', '.webp')):
            image = cv2.imread(file_path)
            if image is not None:
                images.append(image)
            else:
                raise ValueError("Failed to load image.")
        else:
            raise ValueError("Unsupported file format.")
    except Exception as e:
        print(f"❌ Error loading file: {e}")
    
    return images

def preprocess_image(image):
    try:
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image
        
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        thresh = cv2.adaptiveThreshold(blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                       cv2.THRESH_BINARY, 11, 2)
        return thresh
    except Exception as e:
        print(f"⚠ Error in preprocessing: {e}")
        return image
    
def remove_spaces_between_numbers(s):
    return re.sub(r'(?<=\d)\s+(?=\d)', '', s)

def extract_text(file_path):
    images = load_input(file_path)
    extracted_text = ""

    for image in images:
        processed_img = preprocess_image(image)
        text = pytesseract.image_to_string(processed_img)
        extracted_text += text + "\n" 

    extracted_text = remove_spaces_between_numbers(extracted_text)

    return extracted_text.strip() if extracted_text else "No text found"


def extract_key_fields(text):
    invoice_data = {}

    patterns = {
        "Total Amount": r"(Total|Grand Total|Final Amount)[^\d]([₹$€]?\s[\d,]+\.?\d*)",
        "Tax Amount": r"(GST|VAT|Tax|SGST|CGST)[^\d]([₹$€]?\s[\d,]+\.?\d*)",
        "Base Amount": r"(Sub-total|Amount before Tax|Taxable Amount)[^\d]([₹$€]?\s[\d,]+\.?\d*)",
        "Discount": r"(Discount)[^\d]([₹$€]?\s[\d,]+\.?\d*)",
        "Invoice Number": r"(Invoice No|Invoice Number|Bill No)[^\d]*([\w-]+)",
        "Invoice Date": r"(Date|Invoice Date|Bill Date)[^\d]*([\d]{1,2}[/\-][\d]{1,2}[/\-][\d]{2,4})"
    }

    for field, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        invoice_data[field] = match.group(2) if match else "Not Found"

    sentences = sent_tokenize(text)
    sender_info, recipient_info = [], []
    capturing_sender, capturing_recipient = False, False

    for sent in sentences:
        lower_sent = sent.lower()
        if any(keyword in lower_sent for keyword in ["from", "seller", "issued by", "address"]):
            capturing_sender, capturing_recipient = True, False
        elif any(keyword in lower_sent for keyword in ["to", "buyer", "billed to", "bill to", "address"]):
            capturing_sender, capturing_recipient = False, True
        elif capturing_sender:
            sender_info.append(sent)
        elif capturing_recipient:
            recipient_info.append(sent)

    invoice_data["Sender Details"] = " ".join(sender_info) if sender_info else "Not Found"
    invoice_data["Recipient Details"] = " ".join(recipient_info) if recipient_info else "Not Found"

    return invoice_data

def extract_numbers_from_invoice(text):
    pattern = re.findall(
        r"([\w\s]+)\s+(\d+KG|\d+)\s+Rs\s*(\d+\.\d+|\d+)\s+Rs\s*(\d+\.\d+|\d+)\s*\(\d+%\)\s+Rs\s*(\d+\.\d+|\d+)", 
        text
    )
    
    extracted_numbers = []
    for item in pattern:
        extracted_numbers.append({
            "Item": item[0].strip(),
            "Quantity": item[1],
            "Price": float(item[2]),
            "Tax": float(item[3]),
            "Total": float(item[4])
        })

    return extracted_numbers


def save_to_excel(data, filename="invoice_data.xlsx"):
    try:
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Invoice Number", "Invoice Date", "Total Amount", "Tax Amount", "Base Amount", "Sender Details", "Recipient Details"])
            
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

        row_data = [
            data.get("Invoice Number", "Not Found"),
            data.get("Invoice Date", "Not Found"),
            data.get("Total Amount", "Not Found"),
            data.get("Tax Amount", "Not Found"),
            data.get("Base Amount", "Not Found"),
            data.get("Sender Details", "Not Found"),
            data.get("Recipient Details", "Not Found")
        ]
        
        ws.append(row_data)

        item_sheet_name = "Invoice_Items"
        if item_sheet_name in wb.sheetnames:
            ws_items = wb[item_sheet_name]
        else:
            ws_items = wb.create_sheet(item_sheet_name)
            ws_items.append(["Invoice Number", "Item Name", "Quantity", "Price", "Tax", "Total"])

        for item in data.get("Items", []):
            ws_items.append([data.get("Invoice Number", "Not Found"),
                             item["Item"], item["Quantity"], item["Price"], item["Tax"], item["Total"]])

        for ws in [ws, ws_items]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(filename)
        print(f"✅ Data saved to {filename}")
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")

def main(file_path):
    print(f"📄 Processing file: {file_path}")
    extracted_text = extract_text(file_path)
    
    if extracted_text and extracted_text != "No text found":
        print("📝 Extracting key fields...")
        extracted_data = extract_key_fields(extracted_text)
        extracted_data["Items"] = extract_numbers_from_invoice(extracted_text)
        save_to_excel(extracted_data)
        print("✅ Extraction & Save Complete!")
    else:
        print("⚠ No meaningful text extracted. Check image quality.")

if _name_ == "_main_":
    main(file_path)